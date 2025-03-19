import os, csv, json, re, glob, openpyxl, shutil, datetime, shutil
from openpyxl import load_workbook
from openpyxl import Workbook
from ruamel.yaml import YAML

yaml = YAML()

def read_json_file(file_path):
	with open(file_path, "r", encoding="utf-8", errors="ignore") as file:
		data = json.load(file)
		return data

def csv_writer(filename):
	write_file = open(str(filename)+".csv","w", encoding="utf-8", newline="", errors="ignore")
	return csv.writer(write_file)

def csv_reader(filename):
	read_file = open(str(filename)+".csv","r", encoding="utf-8")
	csv_reader = csv.reader(read_file)
	return [a for a in csv_reader]

def error_reader(filename):
	read_file = open(str(filename)+".csv","r", encoding="utf-8")
	csv_reader = csv.reader(read_file)
	return [[a[0] for a in csv_reader], read_file]

def write_csv(dataset):
	# write the scraped data to a .csv file
	with open("job_listings.csv", "w", newline="", encoding="utf8", errors="ignore") as file:
		writer = csv.writer(file)
		writer.writerow(["Title", "Link"])
		writer.writerows(dataset)

def check_errors(dir_path):
	fail_log_file = csv_writer("error_logger")
	for a, b, c in os.walk(dir_path):
		for d in c:
			json_path = str(a+"/"+d).replace("\\\\", "/")
			json_path = re.sub(r'\\', '/', json_path)
			json_data = read_json_file(json_path)
			try:
				if json_data["meta"]["status"].lower() == "error":
					print("walk_directory: json_path: ", json_path)
					fail_log_file.writerow([json_path])
			except:
				pass

def convert_json(dir_path, error_log):
	for a, b, c in os.walk(dir_path):
		#############################
		xl_filename = str(a).replace("\\google_play_data", "")
		xl_filename = str(xl_filename).split("\\")
		xl_filename = xl_filename[len(xl_filename)-1]
		xl_filename = a.replace("\\google_play_data", "")+"/"+xl_filename
		xl_filename = str(xl_filename).replace("\\\\", "/")
		xl_filename = re.sub(r'\\', '/', xl_filename)+".xlsx"

		#############################
		workbook = ""
		#############################
		print("convert_json: xl_filename: ", str(xl_filename))
		if os.path.isfile(xl_filename) == True and ".xml" not in str(xl_filename):
			workbook = load_workbook(filename=xl_filename)
		else:
			workbook = Workbook()
		#############################
		current_dir = a.split("\\")
		current_dir = current_dir[len(current_dir)-1]
		#############################
		if current_dir is not "google_play_data" and "google_play_data" not in str(a) and "google_play_data" not in str(a) and a is not "sw_data" and current_dir not in str(error_log):
			#############################
			for d in c:
				if os.path.isdir("sw_data/"+str(d).replace("xlsx", "")) == False and "google_play_data" not in str(a):
					#############################
					json_path = str(a+"/"+d).replace("\\\\", "/")
					json_path = re.sub(r'\\', '/', json_path)
					#############################
					sheetname = json_path.split("-")
					sheetname = sheetname[len(sheetname)-1]
					sheetname = sheetname.split("_", 1)[1].replace(".json", "")
					############################
					#############################
					if json_path not in error_log and "describe" not in str(json_path) and ".xlsx" not in str(d) and ".xml" not in str(d) and ".json" in str(d):
						dataset = read_json_file(json_path)
						try:
							test_xls(dataset, workbook, sheetname)
						except:
							pass
						#############################
					#############################
		elif "google_play_data" in str(a):
			for d in c:
				if os.path.isdir("sw_data/"+str(d).replace("xlsx", "")) == False:
					#############################
					json_path = str(a+"/"+d).replace("\\\\", "/")
					json_path = re.sub(r'\\', '/', json_path)
					#############################
					sheetname = json_path.split("-")
					sheetname = sheetname[len(sheetname)-1]
					sheetname = sheetname.split("_", 1)[1].replace(".json", "")
					#############################
					#############################
					if json_path not in error_log and "describe" not in str(json_path) and ".xlsx" not in str(d) and ".xml" not in str(d) and ".json" in str(d):
						dataset = read_json_file(json_path)
						gplay_xl_setup_file(dataset, workbook, sheetname)

		workbook.save(xl_filename)
		workbook.close()

def fix_json_string_bs(json_string):
	new_data = re.sub(r'\n\s*(\S+)\s*:', r'\n"\1":', json_string)
	new_data = re.sub(r',\s*}', r'\n}', new_data)
	new_data = re.sub(r'<[^>]*>', '', new_data)
	new_data = json.dumps(new_data)
	new_data = json.loads(new_data)
	########################
	new_data = re.sub(r'[\x00-\x1F\x7F-\x9F]', '', new_data)
	########################
	new_data = json.dumps(yaml.load(new_data))
	new_data = json.loads(new_data)
	return new_data

def gplay_xl_setup_file(dataset, workbook, sheetname):
	#############################  
	#############################
	for a in dataset:
		if dataset[a] != None:
			if a.lower() == "app_overview":
				headers = gplay_get_headers(dataset[a])
				privacy_url = ""
				new_data = dataset["data_safety"]
				if isinstance(new_data, str):
					new_data = fix_json_string_bs(new_data)

				if new_data and new_data["privacyPolicyUrl"]:
					privacy_url = new_data["privacyPolicyUrl"]
				gplay_process_data(dataset[a], workbook, sheetname, headers, a, privacy_url)

			elif a.lower() == "dev_info":
				headers = gplay_get_headers(dataset[a][0])
				gplay_process_data(dataset[a][0], workbook, sheetname, headers, a, "")
			
			elif a.lower() == "data_safety":
				new_data = dataset[a]
				if isinstance(new_data, str) == True:
					new_data = fix_json_string_bs(new_data)
				for b in new_data:
					if len(b) > 0 and b != "privacyPolicyUrl" and len(new_data[b]) > 0:
						headers = gplay_get_headers(new_data[b][0])
						gplay_process_data(new_data[b], workbook, sheetname, headers, b, "")

			elif a.lower() == "permissions":
				if len(dataset[a]) > 0:	
					headers = gplay_get_headers(dataset[a][0])
					gplay_process_data(dataset[a], workbook, sheetname, headers, a, "")

def gplay_process_data(dataset, workbook, sheetname, headers, obj_key, privacy_url):
	existing_sheets = workbook.sheetnames
	if obj_key == "app_overview":
		headers.append("privacyPolicyUrl")
	worksheet = ""
	#############################
	worksheet_title = "gplay_"+str(obj_key)
	#############################
	if len(existing_sheets) == 1 and (existing_sheets[0] == "Sheet1" or existing_sheets[0] == "Sheet"):
		worksheet = workbook.active
		worksheet.title = worksheet_title
	elif worksheet_title not in existing_sheets:
		workbook.create_sheet(worksheet_title)
		worksheet = workbook[worksheet_title]
	else:
		worksheet = workbook[worksheet_title]
	# #############################
	first_column_values = get_first_col_val(worksheet)
	# #############################
	# #############################
	if len(list(worksheet.iter_rows(values_only=True))) <= 0:
		worksheet.append(headers)
	# #############################
	
	if obj_key == "app_overview":
		headers.append("privacyPolicyUrl")
		compiled_data = gplay_get_values(dataset, headers)
		compiled_data.append(privacy_url)
		if compiled_data[0] not in first_column_values:
			worksheet.append(compiled_data)
	
	elif obj_key == "dev_info":
		if dataset["title"] not in first_column_values:
			worksheet.append([dataset["title"], dataset["appId"], dataset["url"], dataset["icon"], dataset["developer"], dataset["currency"], dataset["price"], dataset["free"], dataset["summary"], dataset["scoreText"], dataset["score"]])
	
	elif obj_key == "data_safety" or obj_key == "sharedData" or obj_key == "collectedData":
		for row, stat in enumerate(dataset):
			if stat["data"] not in first_column_values:
				worksheet.append([stat["data"], stat["optional"], stat["purpose"], stat["type"]])

	elif obj_key == "data_safety" or obj_key == "securityPractices":
		for row, stat in enumerate(dataset):
			if stat["practice"] not in first_column_values:
				worksheet.append([stat["practice"], stat["description"]])

	elif obj_key == "permissions":
		for row, stat in enumerate(dataset):
			if stat["permission"] not in first_column_values:
				worksheet.append([stat["permission"], stat["type"]])

def gplay_get_values(data_object, headers):
	new_values = []
	for a in headers:
		if str(a) == "histogram" or str(a) == "screenshots"or str(a) == "comments":
			new_values.append(str(data_object[a]))
		elif str(a) != "privacyPolicyUrl":
			new_values.append(data_object[a])
	return new_values

def gplay_get_headers(data_object):
	headers = data_object.keys()
	return list(headers)

def test_xls(dataset, workbook, sheetname):
	#############################
	attr = {
		"audience_app_demographics_age":{
			"alias": "demo_age",
			"level_1": "demographics_data",
			"level_2": "demographics_values",
			"headers":["date", "age_18_to_24", "age_25_to_34", "age_35_to_44", "age_45_to_54", "age_55_plus"]
		},
		"audience_app_demographics_gender":{
			"alias": "demo_gender",
			"level_1": "demographics_data",
			"level_2": "demographics_values",
			"headers":["date", "male", "female"]
		},
		"downloads_downloads":{
			"alias": "downloads",
			"level_1": "downloads",
			"level_2": "",
			"headers":["date", "value"]
		},
		"retention_day_30_retention":{
			"alias": "retention",
			"level_1": "app_retention",
			"level_2": "retention_days",
			"headers":["date", "retention_day", "retention_value"]
		},
		"affinity_affinity":{
			"alias": "affinity",
			"level_1": "also_used_apps",
			"level_2": "",
			"headers":["application", "affinity"]
		},
		"engagement_average_session_length":{
			"alias": "session_len",
			"level_1": "average_session_length",
			"level_2": "",
			"headers":["date", "value"]
		},
		"engagement_open_rate":{
			"alias": "open_rate",
			"level_1": "open_rate",
			"level_2": "",
			"headers":["date", "value"]
		},
		"engagement_average_monthly_user_sessions":{
			"alias": "monthly_sessions",
			"level_1": "average_monthly_user_sessions",
			"level_2": "",
			"headers":["date", "value"]
		},
		"engagement_average_monthly_user_time_spent":{
			"alias": "monthly_time",
			"level_1": "average_monthly_user_time_spent",
			"level_2": "",
			"headers":["date", "value"]
		},
		"engagement_install_base":{
			"alias": "install_base",
			"level_1": "install_base",
			"level_2": "",
			"headers":["date", "value"]
		},
		"engagement_monthly_active_users":{
			"alias": "mau",
			"level_1": "monthly_active_users",
			"level_2": "",
			"headers":["date", "value"]
		},
		"engagement_total_sessions":{
			"alias": "total_sessions",
			"level_1": "total_sessions",
			"level_2": "",
			"headers":["date", "value"]
		},
		"engagement_total_time":{
			"alias": "total_time",
			"level_1": "total_time",
			"level_2": "",
			"headers":["date", "value"]
		}
	}
	#############################
	lvl_a = ""
	if attr[sheetname]["alias"] == "retention":
		try:
			lvl_a = dataset["app_retention"]
		except:
			pass

		try:
			lvl_a = dataset["retention"]
		except:
			pass

	else:
		lvl_a = dataset[attr[sheetname]["level_1"]]

	headers = attr[sheetname]["headers"]
	#############################
	existing_sheets = workbook.sheetnames
	worksheet = ""
	#############################
	#############################
	if len(existing_sheets) == 1 and (existing_sheets[0] == "Sheet1" or existing_sheets[0] == "Sheet"):
		worksheet = workbook.active
		worksheet.title = attr[sheetname]["alias"]
	elif attr[sheetname]["alias"] not in existing_sheets:
		workbook.create_sheet(attr[sheetname]["alias"])
		worksheet = workbook[attr[sheetname]["alias"]]
	else:
		worksheet = workbook[attr[sheetname]["alias"]]
	#############################
	first_column_values = get_first_col_val(worksheet)
	#############################
	#############################
	if len(list(worksheet.iter_rows(values_only=True))) <= 0:
		worksheet.append(headers)
	#############################
	if attr[sheetname]["level_2"] == "":
		for row, stat in enumerate(lvl_a):
			if attr[sheetname]["alias"] == "affinity" and stat["application"] not in first_column_values:
				worksheet.append([stat["application"], stat["affinity"]])
			elif attr[sheetname]["alias"] != "affinity" and stat["date"] not in first_column_values:
				current_val = stat["value"]
				if current_val == None:
					current_val = 0
				worksheet.append([stat["date"], current_val])
	else:
		for row, stat in enumerate(lvl_a):
			if stat["date"] not in first_column_values:
				if "retention_days" in stat:
					for a in stat["retention_days"]:
						worksheet.append([stat["date"], a["retention_day"], a["retention_value"]])

				elif "demographics_values" in stat:
					worksheet.append([stat["date"]]+demo_stats(stat))

def demo_stats(stat):
	vals = []
	try:
		if stat["demographics_values"]:
			for a in stat["demographics_values"]:
				demo_val = stat["demographics_values"][a]
				if demo_val == None:
					demo_val = 0.00
				vals.append(demo_val)
	except:
		pass

	return vals

def get_first_col_val(worksheet):
	return [a[0] for a in worksheet.values]

def rm_all_xl(dir_path):
	for a, b, c in os.walk(dir_path):
		for d in c:
			if ".xlsx" in str(d):
				try:
					xl_filename = str(a).split("\\")
					xl_filename = xl_filename[len(xl_filename)-1]
					xl_filename = a+"/"+xl_filename
					xl_filename = str(xl_filename).replace("\\\\", "/")
					xl_filename = re.sub(r'\\', '/', xl_filename)+".xlsx"
					xl_filename = "./"+xl_filename
					os.remove(xl_filename)
					print("rm_all_xl: xl_filename: ", xl_filename)
				except:
					pass

def backup_directory(src_dir, error_log):
	# Create the backup directory name using the current date
	today = datetime.date.today()
	backup_dir_name = today.strftime("%m_%d_%Y")
	
	# Create the backup directory path
	backup_dir_path = os.path.join("0_archive", backup_dir_name)
	
	# Create the backup directory if it doesn't already exist
	os.makedirs(backup_dir_path, exist_ok=True)
	
	# Backup the source directory to the backup directory
	shutil.copytree(src_dir, os.path.join(backup_dir_path, os.path.basename(src_dir)))
	
	print(f"Directory {src_dir} has been backed up to {backup_dir_path}.")

def backup_excel_files(src_dir, error_log):
	# Create the backup directory name using the current date
	today = datetime.date.today()
	backup_dir_name = "0_xlsx"
	
	# Create the backup directory path
	backup_dir_path = os.path.join(backup_dir_name, today.strftime("%m_%d_%Y"))
	
	print("backup_excel_files: backup_dir_name: ", backup_dir_name)
	
	# Create the backup directory if it doesn't already exist
	os.makedirs(backup_dir_path, exist_ok=True)
	
	# Loop through the source directory and its subdirectories
	for dirpath, dirnames, filenames in os.walk(src_dir):
		for filename in filenames:
			# Only copy .xlsx files
			if filename.endswith('.xlsx') and str(filename).replace(".xlsx", "") not in error_log:
				# Get the full source file path
				src_file_path = os.path.join(dirpath, filename)
				
				# Get the destination file path
				dst_file_path = os.path.join(backup_dir_path, filename)
				
				
				print("backup_excel_files: dirpath: ", dirpath)
				print("backup_excel_files: dirnames: ", dirnames)
				print("backup_excel_files: filename: ", filename)
				print("backup_excel_files: filenames: ", filenames)
				print("backup_excel_files: src_file_path: ", src_file_path)
				print("backup_excel_files: dst_file_path: ", dst_file_path)
				print("*"*50)
				print("*"*50)
				

				# Copy the file to the destination directory
				shutil.copy2(src_file_path, dst_file_path)
	
	print(f"Excel files in {src_dir} have been backed up to {backup_dir_path}.")

def copy_xlsx_files(directory_path):
	for root, dirs, files in os.walk(directory_path):
		for file in files:
			if file.endswith(".xlsx"):
				source_path = os.path.join(root, file).replace("\\", "/")
				destination_path = source_path.replace(directory_path+"/", "").replace("\\", "/")
				print(destination_path)
				destination_dir = os.path.dirname(destination_path)
				os.makedirs(destination_dir, exist_ok=True)
				shutil.copy2(source_path, destination_path)

def init():
	#############################
	error_log = error_reader("error_logger")
	error_log[1].close()
	print(error_log[0])
	#############################
	backup_excel_files("sw_data", error_log[0])
	#############################

init()