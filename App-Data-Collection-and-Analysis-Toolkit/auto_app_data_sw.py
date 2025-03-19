import os, openpyxl, requests, time, random, pickle, csv, glob
from datetime import datetime
from bs4 import BeautifulSoup as soup
from openpyxl import load_workbook
from openpyxl import Workbook

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver import DesiredCapabilities
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys

from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service

options = Options()
options.add_experimental_option("prefs", { "download.default_directory": r"C:\Users\user\Desktop\0__0\1_user\_similarweb\0_sw\0_selenium_app_data_apple"})

driver = webdriver.Chrome(service = Service("./chromedriver.exe"), options=options)

driver.set_window_position(800, 0)
driver.set_window_size(850, 600)

actions = ActionChains(driver) 

# Replace with your credentials
username = "your_email@example.com"
password = "your_password"

sources = ["all traffic", "desktop", "mobile web"]

duration_checked = False

def save_cookies():
	pickle.dump( driver.get_cookies() , open("cookies.pkl","wb"))

def load_cookies():
	cookies = pickle.load(open("cookies.pkl", "rb"))
	for cookie in cookies:
		driver.add_cookie(cookie)

def click_element(selector):
	print("auto_sw: click_element: selector: ", selector)
	driver.find_element(By.CSS_SELECTOR, selector).click()
	time.sleep(2)

def key_entry(key_value):
	actions = ActionChains(driver) 
	actions.send_keys(key_value)
	actions.perform()

def type_input(string_value):
	for character in string_value:
		key_speed = random.uniform(1/6.8, 1/8.0)
		if character == "." or character == "@" or character.isdigit():
			key_speed = random.uniform(1/5.04, 1/6.5)
		print("auto_sw: login: key_speed: ", key_speed)
		print("auto_sw: login: character: ", character)
		key_entry(character)
		time.sleep(key_speed)

def login():
	###################################
	click_element("#input-email")
	type_input(username)
	key_entry(Keys.TAB)
	type_input(password)
	key_entry(Keys.ENTER)
	###################################

def get_page(url):
	driver.get(url)
	time.sleep(1)
	a = driver.execute_script("return document.getElementsByTagName('html')[0].innerHTML")
	return a

def wait_main_page():
	main_loaded = False
	while main_loaded == False:
		greeting_text = driver.find_elements(By.CSS_SELECTOR, "div")
		for div in greeting_text:
			try:
				if div.text == "Hi User" or "Hi, Welcome to Similarweb" in div.text:
					main_loaded = True
					print("auto_sw: wait_main_page: div.text: ", div.text)
					time.sleep(1)
			except:
				pass

def get_files(url):
	driver.execute_script("window.open('');")
	time.sleep(0.5)
	driver.switch_to.window(driver.window_handles[1])
	time.sleep(0.5)
	print("auto_sw: get_files: url: ", str(url))
	driver.get(url)
	time.sleep(0.5)
	driver.close()
	driver.switch_to.window(driver.window_handles[0])

def get_ids(filename):
	a = open(filename+".csv", "r", encoding="utf-8", newline="", errors="ignore")
	b = csv.reader(a)
	c = [c for c in b]
	a.close()
	return c

def create_folder(dir_path):
	if not os.path.exists(str(dir_path)):
		os.makedirs(dir_path)

def get_date_name():
	###################################
	current_month 	= str(datetime.now().month)
	if len(current_month) == 1:
		current_month = "0"+str(current_month)
	###################################
	current_day 	= str(datetime.now().day)
	if len(current_day) == 1:
		current_day = "0"+str(current_day)
	###################################
	current_year 	= str(datetime.now().year)
	###################################
	return str(current_month+current_day+current_year)
	###################################

def get_worksheet_rows(existing_worksheet):
	all_rows = []

	for count, row in enumerate(existing_worksheet):
		current_row = []
		for cell in row:
			current_row.append(cell.value)
		all_rows.append(current_row)
	return all_rows

def get_last_filename_and_rename(save_folder, app_id, engagement_type):
	###################################
	files = glob.glob(save_folder + '/*')
	files = [check_file for check_file in files if not os.path.isdir(check_file)]
	if len(files) > 0:
		try:
			###################################
			max_file 		= max(files, key=os.path.getctime)
			filename 		= max_file.split("\\")[-1].split(".")[0]
			edit_filename 	= filename.replace(" ", "")
			###################################
			if get_date_name() not in filename and str(filename)[:10] != "Engagement":
				new_path = max_file.replace(filename, str(app_id).replace(".", "").lower()+"_"+str(edit_filename)+"_"+get_date_name()).replace(" ", "")
				###################################
				print("get_last_filename_and_rename: app_id: ", str(app_id))
				print("get_last_filename_and_rename: max_file: ", str(max_file))
				print("get_last_filename_and_rename: filename: ", str(filename))
				print("get_last_filename_and_rename: edit_filename: ", str(edit_filename))
				print("get_last_filename_and_rename: new_path: ", str(new_path))
				###################################
				print("*"*50)
				print("*"*50)
				os.rename(max_file, new_path)
		except:
			pass
		###################################

def catch_engagement(save_folder, app_id):
	###################################
	files = glob.glob(save_folder + '/*')
	files = [check_file for check_file in files if not os.path.isdir(check_file)]

	if len(files) > 0:
		for a in files:
			filename = a.split("\\")[-1]
			if str(filename)[:10] == "Engagement":
				###################################
				existing_workbook 		= load_workbook(filename=a)
				existing_sheets 		= existing_workbook.sheetnames
				current_worksheet 		= existing_workbook[existing_sheets[1]]
				current_rows 			= get_worksheet_rows(current_worksheet)
				###################################
				filename_string = filename.split(".")[0]
				###################################
				engagement_type = ""
				if "Total Time" in current_rows[0][1]:
					engagement_type = "_EngagementSessions_"
				elif "Monthly Active Users" in current_rows[0][1]:
					engagement_type = "_EngagementOpenRate_"
				
				###################################
				new_path = a.replace(filename_string, str(app_id).replace(".", "").lower()+str(engagement_type)+get_date_name())
				###################################
				print("catch_engagement: a: ", str(a))
				print("catch_engagement: new_path: ", str(new_path))
				###################################
				print("*"*50)
				print("*"*50)
				try:
					os.rename(a, new_path)
				except:
					pass

def download_app_files_apple(app_list, downloaded_ids):
	###################################
	sleep_from = 0.3
	sleep_to = 1.25
	###################################
	root_dir_path 	= r"C:\Users\user\Desktop\0__0\1_user\_similarweb\0_sw\0_selenium_app_data_apple"
	###################################
	for a, b in enumerate(app_list):
		###################################
		if a > 0:
			###################################
			app_dir_path 	= "\\"+str(b[0])
			complete_path 	= root_dir_path + app_dir_path
			print("download_app_files_android: complete_path: ", str(complete_path))
			print("row count: ", str(a))
			###################################
			app_id 					= str(b[0]).replace(" ", "")
			print("app_id: ", str(app_id))
			download_entry_name  	= str(b[0]).replace(" ", "").replace(".", "").lower()
			###################################
			###################################
			if download_entry_name not in downloaded_ids and app_id != "com.example.app1" and "example.app2" not in str(app_id):
				###################################
				get_files("https://pro.similarweb.com/api/AppAnnie/InstallBase/Graph/Excel?%23&country=840&device=Combined&from=2022%7C02%7C01&keys="+str(app_id)+"&store=apple&to=2023%7C04%7C30")
				time.sleep(0.5)
				get_last_filename_and_rename(root_dir_path, app_id, "")
				time.sleep(random.uniform(sleep_from, sleep_to))
				###################################
				get_files("https://pro.similarweb.com/api/AppAnnie/InstallBase/Delta/Excel?%23&country=840&device=total&from=2022%7C02%7C01&keys="+str(app_id)+"&store=apple&to=2023%7C04%7C30")
				time.sleep(0.5)
				get_last_filename_and_rename(root_dir_path, app_id, "")
				time.sleep(random.uniform(sleep_from, sleep_to))
				###################################
				get_files("https://pro.similarweb.com/api/AppAnnie/Engagement/Mau/Excel?country=840&from=2022%7C02%7C01&to=2023%7C04%7C30&keys="+str(app_id)+"&store=apple&timeGranularity=Monthly&isWindow=false&device=Combined")
				time.sleep(0.5)
				get_last_filename_and_rename(root_dir_path, app_id, "open")
				time.sleep(random.uniform(sleep_from, sleep_to))
				###################################
				get_files("https://pro.similarweb.com/api/AppAnnie/Engagement/Excel?country=840&from=2022%7C02%7C01&to=2023%7C04%7C30&keys="+str(app_id)+"&store=apple&timeGranularity=Monthly&isWindow=false&device=Combined")
				time.sleep(0.5)
				get_last_filename_and_rename(root_dir_path, app_id, "time")
				time.sleep(random.uniform(sleep_from, sleep_to))
				###################################
				get_files("https://pro.similarweb.com/api/AppAnnie/Retention/Excel?%23&country=840&device=Total&from=2022%7C01%7C01&keys="+str(app_id)+"&store=apple&to=2023%7C03%7C31")
				time.sleep(0.5)
				get_last_filename_and_rename(root_dir_path, app_id, "")
				time.sleep(random.uniform(sleep_from, sleep_to))
				###################################
				get_files("https://pro.similarweb.com/widgetApi/AppAnnie/AudienceInterests/Excel?asc=false&country=840&device=iPhone&from=2022%7C02%7C01&keys="+str(app_id)+"&sort=Affinity&store=apple&to=2023%7C04%7C30")
				time.sleep(0.5)
				get_last_filename_and_rename(root_dir_path, app_id, "")
				time.sleep(random.uniform(sleep_from, sleep_to))
				#################################
				get_files("https://pro.similarweb.com/widgetApi/AppRanks/AppRanksHistory/Excel?keys="+str(app_id)+"&store=Apple&timeGranularity=Daily&to=2023%7C05%7C18&isWindow=false&appmode=Top%20Free&includeSubDomains=true&from=2020%7C01%7C01&country=840&isDaily=true")
				time.sleep(0.5)
				get_last_filename_and_rename(root_dir_path, app_id, "")
				time.sleep(random.uniform(sleep_from, sleep_to))
				##################################
				catch_engagement(root_dir_path, app_id)
				##################################
				print("*"*50)
				print("*"*50)
				print("*"*50)
				##################################
			else:
				print("IN: ", str(b))

def write_csv(filename):
	a = open(filename+".csv", "w", encoding="utf-8", newline="", errors="ignore")
	b = csv.writer(a)
	return [b, a]

def get_app_ids_from_filenames():
	###################################
	root_dir_path 	= r"C:\Users\user\Desktop\0__0\1_user\_similarweb\0_sw\0_selenium_app_data_apple"
	###################################
	app_list = []
	###################################
	for a, b in enumerate(os.listdir(root_dir_path)):
		###################################
		print("get_app_ids_from_filenames: b: ", str(b))
		###################################
		app_id 			= str(b).split("_")[0]
		app_id_alt 		= str(b).split("_AppsDemographics")[0]
		app_id_alt 		= str(app_id_alt).split("_EngagementOpenRate")[0]
		app_id_alt 		= str(app_id_alt).split("_InstallBase")[0]
		app_id_alt 		= str(app_id_alt).split("_InstallBaseDelta")[0]
		app_id_alt 		= str(app_id_alt).split("_Retention")[0]
		app_id_alt 		= str(app_id_alt).split("_StoreDownloads")[0]
		###################################
		if app_id not in app_list and app_id_alt not in app_list:
			app_list.append(app_id)
			app_list.append(app_id_alt)
		###################################
	return app_list
	###################################

def compare_ids_to_files(app_list, downloaded_ids):
	###################################
	csv_downloaded_log 	= write_csv("auto_app_data_sw_log")
	close_log 			= csv_downloaded_log[1]
	csv_downloaded_log 	= csv_downloaded_log[0]
	root_dir_path 		= "0_selenium_app_data"
	###################################
	for a, b in enumerate(app_list):
		###################################
		if a > 0:
			###################################
			app_id 			= str(b[0]).replace(" ", "").replace(".", "").lower()
			###################################
			if app_id in downloaded_ids:
				###################################
				print("compare_ids_to_files: b: ", str(b))
	close_log.close()

def init_headless():
	###################################
	id_csv_path 	= "apple_id"
	app_list 		= get_ids(id_csv_path)
	downloaded_ids 	= []
	print("init_headless: downloaded_ids: ", str(app_list))
	print("init_headless: downloaded_ids length: ", str(len(downloaded_ids)))
	###################################
	###################################
	login_url = "https://pro.similarweb.com/"
	get_page(login_url)
	login()
	wait_main_page()
	time.sleep(1)
	###################################
	###################################
	download_app_files_apple(app_list, downloaded_ids)
	###################################

init_headless()