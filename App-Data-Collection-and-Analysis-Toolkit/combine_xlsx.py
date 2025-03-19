import os, csv, re, openpyxl, datetime, json, glob
from openpyxl import load_workbook
from openpyxl import Workbook
import fnmatch


def csv_writer(filename):
	write_file = open(str(filename)+".csv","w", encoding="utf-8", newline="", errors="ignore")
	return csv.writer(write_file)

def check_columns(existing_sheets):
	must_contain_columns = ['demo_age', 'demo_gender', 'downloads', 'monthly_sessions', 'monthly_time', 'session_len', 'install_base', 'mau', 'open_rate', 'total_sessions', 'total_time']
	for a in must_contain_columns:
		if str(a) not in existing_sheets:
			return False
	return True

def check_gplay_data(existing_sheets):
	for a in existing_sheets:
		if "gplay" in str(a):
			return True
	return False


def organize_col_values(row_data, sheet_type, temp_object):
	if sheet_type == "demo_age":
		temp_object["demo_age_18_to_24"] = row_data[1]
		temp_object["demo_age_25_to_34"] = row_data[2]
		temp_object["demo_age_35_to_44"] = row_data[3]
		temp_object["demo_age_45_to_54"] = row_data[4]
		temp_object["demo_age_55_plus"] = row_data[5]
	
	elif sheet_type == "demo_gender":
		temp_object["demo_gender_male"] = row_data[1]
		temp_object["demo_gender_female"] = row_data[2]
	
	else:
		temp_object[sheet_type] = row_data[1]

	return temp_object 


def find_file(file_name, search_path):
	result = []
	for root, dirs, files in os.walk(search_path):
		for name in files:
			if fnmatch.fnmatch(name, file_name):
				result.append(os.path.join(root, name))
	return result


def read_csv(filepath):
	a = open(filepath, "r")
	b = csv.reader(a)
	return [c for c in b]

def edit_categories(csv_categories, combined_data):
	for a in combined_data:
		for b in csv_categories:
			if str(a["app_name"]).lower().replace(".", "") == str(b[0]).lower().replace(".", ""):
				a["category"] = b[1]
	return combined_data


############################################################################################################
############################################################################################################


def get_retention_rows(all_rows):
	new_rows = []
	new_arr = []
	current_date = None

	for a in all_rows:
		if a[0] == current_date:
			new_arr.append(a[2])
		else:
			if new_arr:
				new_rows.append(new_arr)
			new_arr = [a[0], a[2]]
			current_date = a[0]

	if new_arr:
		new_rows.append(new_arr)
	
	return new_rows


def fix_csv_categories(combined_filepath, ids_filepath):
	combined_csv = read_csv(combined_filepath)
	ids_csv = read_csv(ids_filepath)
	for a in combined_csv:
		for b in ids_csv:
			if str(a[0]).lower().replace(".", "") == str(b[0]).lower().replace(".", ""):
				a[1] = b[1]
	write_csv = csv_writer("0_xlsx/0_combined_xlsx//0_combined_data")
	write_csv.writerows(combined_csv)


def get_worksheet_rows(existing_worksheet):
	all_rows = []

	for count, row in enumerate(existing_worksheet):
		if count != 0:	
			current_row = []
			for cell in row:
				current_row.append(cell.value)
			all_rows.append(current_row)
	return all_rows


def test_sheet(app_name, existing_workbook, existing_sheets):
	storage_array = []

	headers = ["retention", "demo_age", "demo_gender", "downloads", "monthly_sessions", "monthly_time", "session_len", "install_base", "mau", "open_rate", "total_sessions", "total_time"]
	
	for sheet_name in existing_sheets:
		if sheet_name != "affinity" and "gplay" not in sheet_name:
			worksheet = existing_workbook[sheet_name]
			all_rows = get_worksheet_rows(worksheet)

			if "retention" in str(sheet_name) and len(all_rows) > 1 and len(all_rows[0]) > 2:
				all_rows = get_retention_rows(all_rows)

			for x, row in enumerate(all_rows):
				date = row[0]

				# Check if a row with this date and app_name already exists in storage_array
				index = next((i for i, obj in enumerate(storage_array) if obj["date"] == date and obj["app_name"] == app_name), -1)
				
				# If the row doesn't exist yet, create a new one and add it to storage_array
				if index == -1:
					temp_object = {
						"app_name": app_name,
						"category": "",
						"date": date,
						"retention":"",
						"retention_day_0": "",
						"retention_day_1": "",
						"retention_day_2": "",
						"retention_day_3": "",
						"retention_day_4": "",
						"retention_day_5": "",
						"retention_day_7": "",
						"retention_day_14": "",
						"retention_day_30": "",
						"demo_age_18_to_24": "",
						"demo_age_25_to_34": "",
						"demo_age_35_to_44": "",
						"demo_age_45_to_54": "",
						"demo_age_55_plus": "",
						"demo_gender_male": "",
						"demo_gender_female": "",
						"downloads": "",
						"monthly_sessions": "",
						"monthly_time": "",
						"session_len": "",
						"install_base": "",
						"mau": "",
						"open_rate": "",
						"total_sessions": "",
						"total_time": "",
						"google_play_data": ""
					}
					storage_array.append(temp_object)
					index = len(storage_array) - 1

				# Update the existing row in storage_array with the new data
				if sheet_name == "demo_age":
					storage_array[index]["demo_age_18_to_24"] = row[1]
					storage_array[index]["demo_age_25_to_34"] = row[2]
					storage_array[index]["demo_age_35_to_44"] = row[3]
					storage_array[index]["demo_age_45_to_54"] = row[4]
					storage_array[index]["demo_age_55_plus"] = row[5]
				elif sheet_name == "demo_gender":
					storage_array[index]["demo_gender_male"] = row[1]
					storage_array[index]["demo_gender_female"] = row[2]
				elif "retention" in sheet_name and len(row) > 2:
					storage_array[index]["retention_day_0"] = row[1]
					storage_array[index]["retention_day_1"] = row[2]
					storage_array[index]["retention_day_2"] = row[3]
					storage_array[index]["retention_day_3"] = row[4]
					storage_array[index]["retention_day_4"] = row[5]
					storage_array[index]["retention_day_5"] = row[6]
					storage_array[index]["retention_day_7"] = row[7]
					storage_array[index]["retention_day_14"] = row[8]
					storage_array[index]["retention_day_30"] = row[9]
				else:
					storage_array[index][sheet_name] = row[1]

			# Update the "google_play_data" field for all rows in storage_array once we've processed all the rows in the current sheet
			for obj in storage_array:
				obj["google_play_data"] = check_gplay_data(existing_sheets)

	return storage_array


############################################################################################################
############################################################################################################


def combine_xlsx(xlxs_dir_and_date_subdir_path, destination_path):
	#############################
	os.makedirs(destination_path, exist_ok=True)
	#############################
	dir_path = xlxs_dir_and_date_subdir_path
	#############################
	qualified_count = 0 
	#############################

	headers = ["app_name", "category", "date", "retention", "retention_day_0", "retention_day_1", "retention_day_2", "retention_day_3", "retention_day_4", "retention_day_5", "retention_day_7", "retention_day_14", "retention_day_30", "demo_age_18_to_24", "demo_age_25_to_34", "demo_age_35_to_44", "demo_age_45_to_54", "demo_age_55_plus", "demo_gender_male", "demo_gender_female", "downloads", "monthly_sessions", "monthly_time", "session_len", "install_base", "mau", "open_rate", "total_sessions", "total_time", "google_play_data"]
	write_csv = csv_writer(destination_path+"/0_combined_data")
	write_csv.writerow(headers)
	#############################
	csv_read_path = "./add_ids_0328.csv"
	read_csv_categories = read_csv(csv_read_path)
	#############################
	#############################
	for a, b, c in os.walk(dir_path):
		#############################
		for d in c:
			if d.endswith('.xlsx'):
				#############################
				
				#############################
				src_file_path = os.path.join(dir_path, d)
				app_name = str(d).replace(".xlsx", "")
				#############################
				#############################
				existing_workbook = load_workbook(filename=src_file_path)
				existing_sheets = existing_workbook.sheetnames
				file_qualifies = check_columns(existing_sheets)
				#############################
				if "combgsocialcardmaker" in str(d):
					print("d: ", str(d))
					print("app_name: ", str(app_name))
					print("src_file_path: ", str(src_file_path))
					print("existing_workbook: ", str(existing_workbook))
					print("existing_sheets: ", str(existing_sheets))
					print("file_qualifies: ", str(file_qualifies))
				#############################
				if file_qualifies == True:
					app_data = test_sheet(app_name, existing_workbook, existing_sheets)
					app_data = edit_categories(read_csv_categories, app_data)
					for z in app_data:
						new_row = [z[x] for x in z]
						write_csv.writerow(new_row)
					qualified_count = qualified_count+1

############################################################################################################
############################################################################################################


def selenium_test_sheet(app_name, existing_workbook, existing_sheets):
	###################################
	storage_array = []
	###################################
	for sheet_name in existing_sheets:
		if sheet_name != "affinity" and "gplay" not in sheet_name:
			###################################
			if "Report Details" in str(sheet_name):
				continue
			###################################
			worksheet = existing_workbook[sheet_name]
			all_rows = get_worksheet_rows(worksheet)
			###################################
			###################################
			for x, row in enumerate(all_rows):
				###################################
				###################################
				date = row[0]
				###################################
				# Check if a row with this date and app_name already exists in storage_array
				index = next((i for i, obj in enumerate(storage_array) if obj["date"] == date and obj["app_name"] == app_name), -1)
				###################################
				# If the row doesn't exist yet, create a new one and add it to storage_array
				###################################
				if index == -1:
					###################################
					temp_object = {
						"app_name": app_name,
						"category": "",
						"date": date,
						"retention":"",
						"retention_day_0": "",
						"retention_day_1": "",
						"retention_day_2": "",
						"retention_day_3": "",
						"retention_day_4": "",
						"retention_day_5": "",
						"retention_day_7": "",
						"retention_day_14": "",
						"retention_day_30": "",
						"demo_age_18_to_24": "",
						"demo_age_25_to_34": "",
						"demo_age_35_to_44": "",
						"demo_age_45_to_54": "",
						"demo_age_55_plus": "",
						"demo_gender_male": "",
						"demo_gender_female": "",
						"downloads": "",
						"monthly_sessions": "",
						"monthly_time": "",
						"session_len": "",
						"install_base": "",
						"mau": "",
						"open_rate": "",
						"total_sessions": "",
						"total_time": "",
						"google_play_data": ""
					}
					storage_array.append(temp_object)
					index = len(storage_array) - 1
					###################################
				###################################


				# Update the existing row in storage_array with the new data
				if sheet_name == "demo_age":
					storage_array[index]["demo_age_18_to_24"] = row[1]
					storage_array[index]["demo_age_25_to_34"] = row[2]
					storage_array[index]["demo_age_35_to_44"] = row[3]
					storage_array[index]["demo_age_45_to_54"] = row[4]
					storage_array[index]["demo_age_55_plus"] = row[5]
				elif sheet_name == "demo_gender":
					storage_array[index]["demo_gender_male"] = row[1]
					storage_array[index]["demo_gender_female"] = row[2]
				elif "retention" in sheet_name and len(row) > 2:
					storage_array[index]["retention_day_0"] = row[1]
					storage_array[index]["retention_day_1"] = row[2]
					storage_array[index]["retention_day_2"] = row[3]
					storage_array[index]["retention_day_3"] = row[4]
					storage_array[index]["retention_day_4"] = row[5]
					storage_array[index]["retention_day_5"] = row[6]
					storage_array[index]["retention_day_7"] = row[7]
					storage_array[index]["retention_day_14"] = row[8]
					storage_array[index]["retention_day_30"] = row[9]
				else:
					storage_array[index][sheet_name] = row[1]

			# Update the "google_play_data" field for all rows in storage_array once we've processed all the rows in the current sheet
			for obj in storage_array:
				obj["google_play_data"] = check_gplay_data(existing_sheets)

	return storage_array


def get_files_with_name(dir_path, app_name):
	matching_files = glob.glob(os.path.join(dir_path, f"*{app_name}*"))
	return matching_files


def strip_app_name(filename):
	app_name = ""
	for metric in ["_AppsDemographics", "_AppRanksHistory","_StoreDownloads", "_Retention", "_InstallBaseDelta","_InstallBase", "_EngagementOpenRate", "_EngagementSessions"]:
		if metric in filename:
			app_name = str(filename).split(metric)[0]
			break
	return app_name


def strip_metric(filename):
	metric_type = ""
	for metric in ["AppsDemographics", "AppRanksHistory", "StoreDownloads", "Retention", "InstallBaseDelta", "InstallBase", "EngagementOpenRate", "EngagementSessions"]:
		if metric in filename:
			metric_type = metric
			break
	return metric_type


def get_all_file_names(dir_path):
	all_files = glob.glob(os.path.join(dir_path, "*"))
	all_file_names = []
	for file in all_files:
		all_file_names.append(os.path.basename(file))
	return all_file_names


def get_col_header(existing_worksheet, sheet_name, app_name, metric_type):
	all_rows = []
	for count, row in enumerate(existing_worksheet):
		if count == 0:	
			current_row = []

			for cell in row:
				if "AppsDemographics" in str(metric_type):
					current_row.append(cell.value)
				elif app_name not in sheet_name.replace(".", "").lower():
					current_row.append(cell.value.replace(" ", "_").lower()+"__"+sheet_name.lower())
				else:
					current_row.append(cell.value.replace(" ", "_").lower()+"__"+sheet_name.lower())

			all_rows = current_row
			break

	return all_rows


def get_all_apple_app_rank_category_headers(dir_path):
	all_rank_files = get_files_with_name(dir_path, "AppRanksHistory")
	all_rank_headers = []
	for a in all_rank_files:
		metric_type = strip_metric(a)
		existing_workbook = load_workbook(filename=a)
		existing_sheets = existing_workbook.sheetnames
		app_name = str(a.split("\\")[1].split("_")[0])
		#############################
		for sheet_count, sheet_name in enumerate(existing_sheets):
			###################################
			if "Report Details" in str(sheet_name):
				continue
			if sheet_count is not 1 and metric_type == "AppsDemographics":
				continue
			#############################
			worksheet = existing_workbook[sheet_name]
			col_header = get_col_header(worksheet, sheet_name, app_name, metric_type)
			for b in col_header:
				if b not in all_rank_headers:
					print(b)
					all_rank_headers.append(b)


def selenium_combine_xlsx(dir_path, destination_path):
	#############################
	os.makedirs(destination_path, exist_ok=True)
	#############################
	logged_ids = []
	headers = ["app_name", "date", "demo_gender_male", "demo_gender_female", "demo_age_18_to_24", "demo_age_25_to_34", "demo_age_35_to_44", "demo_age_45_to_54", "demo_age_55_plus", "mau_phone", "open_rate_phone", "store_downloads_phone", "sessions_total_time_minutes_phone", "sessions_total_phone", "sessions_avg_per_user_phone", "sessions_avg_time_per_user_minutes_phone", "sessions_avg_length_per_user_minutes_phone", "mau_tablet", "open_rate_tablet", "store_downloads_tablet", "sessions_total_time_minutes_tablet", "sessions_total_tablet", "sessions_avg_per_user_tablet", "sessions_avg_time_per_user_minutes_tablet", "sessions_avg_length_per_user_minutes_tablet", "store_downloads_total", "install_base_delta", "install_base_delta_percentage", "retention_day_0", "retention_day_1", "retention_day_2", "retention_day_3", "retention_day_4", "retention_day_5", "retention_day_7", "retention_day_14", "retention_day_30", "install_base", "health_&_fitness__app_ranking", "entertainment__app_ranking", "finance__app_ranking", "travel__app_ranking", "all__app_ranking", "education__app_ranking", "shopping__app_ranking", "weather__app_ranking", "sports__app_ranking", "business__app_ranking", "medical__app_ranking", "social_networking__app_ranking", "games__app_ranking", "games_/_arcade__app_ranking", "games_/_simulation__app_ranking", "games_/_puzzle__app_ranking", "games_/_role_playing__app_ranking", "navigation__app_ranking", "games_/_card__app_ranking", "games_/_casino__app_ranking", "games_/_adventure__app_ranking", "games_/_strategy__app_ranking", "lifestyle__app_ranking", "games_/_action__app_ranking", "reference__app_ranking", "games_/_sports__app_ranking", "games_/_board__app_ranking", "games_/_word__app_ranking", "games_/_family__app_ranking", "productivity__app_ranking", "games_/_racing__app_ranking", "newsstand__app_ranking", "utilities__app_ranking", "photo_&_video__app_ranking", "games_/_music__app_ranking", "news__app_ranking", "books__app_ranking"]
	write_csv = csv_writer(destination_path+"/0_combined_data_selenium_apple")
	write_csv.writerow(headers)
	#############################
	#############################
	for x, a in enumerate(os.listdir(dir_path)):
		#############################
		src_file_path = os.path.join(dir_path, a)
		app_name = strip_app_name(a)
		#############################
		if len(str(app_name)) > 1 and app_name not in logged_ids:
			#############################
			logged_ids.append(app_name)
			print("file_number: x: ", str(x))
			print("app_name: ", str(app_name))
			files_with_name = get_files_with_name(dir_path, app_name)
			#############################
			storage_array = []
			#############################
			for b in files_with_name:
				#############################
				#############################
				metric_type = strip_metric(b)
				existing_workbook = load_workbook(filename=b)
				existing_sheets = existing_workbook.sheetnames
				#############################
				for sheet_count, sheet_name in enumerate(existing_sheets):
					###################################
					if "Report Details" in str(sheet_name):
						continue
					if sheet_count is not 1 and metric_type == "AppsDemographics":
						continue
					#############################
					worksheet = existing_workbook[sheet_name]
					all_rows = get_worksheet_rows(worksheet)
					col_header = get_col_header(worksheet, sheet_name, app_name, metric_type)
					#############################
					for x, row in enumerate(all_rows):
						#############################
						date = row[0]
						#############################
						index = next((i for i, obj in enumerate(storage_array) if obj["date"] == date and obj["app_name"] == app_name), -1)
						#############################
						if index == -1:
							#############################
							temp_object = {
								"app_name": app_name,
								"date": date,
								"demo_gender_male": "",
								"demo_gender_female": "",
								"demo_age_18_to_24": "",
								"demo_age_25_to_34": "",
								"demo_age_35_to_44": "",
								"demo_age_45_to_54": "",
								"demo_age_55_plus": "",
								"mau_phone": "",
								"mau_phone": "",
								"open_rate_phone": "",
								"store_downloads_phone": "",
								"sessions_total_time_minutes_phone": "",
								"sessions_total_phone": "",
								"sessions_avg_per_user_phone": "",
								"sessions_avg_time_per_user_minutes_phone": "",
								"sessions_avg_length_per_user_minutes_phone": "",
								"mau_tablet": "",
								"open_rate_tablet": "",
								"store_downloads_tablet": "",
								"sessions_total_time_minutes_tablet": "",
								"sessions_total_tablet": "",
								"sessions_avg_per_user_tablet": "",
								"sessions_avg_time_per_user_minutes_tablet": "",
								"sessions_avg_length_per_user_minutes_tablet": "",
								"store_downloads_total": "",
								"install_base_delta": "",
								"install_base_delta_percentage": "",
								"retention_day_0": "",
								"retention_day_1": "",
								"retention_day_2": "",
								"retention_day_3": "",
								"retention_day_4": "",
								"retention_day_5": "",
								"retention_day_7": "",
								"retention_day_14": "",
								"retention_day_30": "",
								"install_base": "",
								"health_&_fitness__app_ranking":"",
								"entertainment__app_ranking":"",
								"finance__app_ranking":"",
								"travel__app_ranking":"",
								"all__app_ranking":"",
								"education__app_ranking":"",
								"shopping__app_ranking":"",
								"weather__app_ranking":"",
								"sports__app_ranking":"",
								"business__app_ranking":"",
								"medical__app_ranking":"",
								"social_networking__app_ranking":"",
								"games__app_ranking":"",
								"games_/_arcade__app_ranking":"",
								"games_/_simulation__app_ranking":"",
								"games_/_puzzle__app_ranking":"",
								"games_/_role_playing__app_ranking":"",
								"navigation__app_ranking":"",
								"games_/_card__app_ranking":"",
								"games_/_casino__app_ranking":"",
								"games_/_adventure__app_ranking":"",
								"games_/_strategy__app_ranking":"",
								"lifestyle__app_ranking":"",
								"games_/_action__app_ranking":"",
								"reference__app_ranking":"",
								"games_/_sports__app_ranking":"",
								"games_/_board__app_ranking":"",
								"games_/_word__app_ranking":"",
								"games_/_family__app_ranking":"",
								"productivity__app_ranking":"",
								"games_/_racing__app_ranking":"",
								"newsstand__app_ranking":"",
								"utilities__app_ranking":"",
								"photo_&_video__app_ranking":"",
								"games_/_music__app_ranking":"",
								"news__app_ranking":"",
								"books__app_ranking":""
							}
							
							#############################
							storage_array.append(temp_object)
							index = len(storage_array) - 1
						###################################
						if "AppsDemographics" in str(b):
							storage_array[index]["demo_gender_male"] = row[1]
							storage_array[index]["demo_gender_female"] = row[2]
							storage_array[index]["demo_age_18_to_24"] = row[3]
							storage_array[index]["demo_age_25_to_34"] = row[4]
							storage_array[index]["demo_age_35_to_44"] = row[5]
							storage_array[index]["demo_age_45_to_54"] = row[6]
							storage_array[index]["demo_age_55_plus"] = row[7]


						elif "EngagementOpenRate" in str(b) and sheet_count == 1:
							storage_array[index]["mau_phone"] = row[1]
							storage_array[index]["open_rate_phone"] = row[2]
						elif "EngagementOpenRate" in str(b) and sheet_count == 2:
							storage_array[index]["mau_tablet"] = row[1]
							storage_array[index]["open_rate_tablet"] = row[2]

						
						
						elif "AppRanksHistory" in str(b) and sheet_count == 1:
							if len(col_header) > 1:
								for c, d in enumerate(col_header):
									if c > 0:
										storage_array[index][d] = row[c]


						elif "EngagementSessions" in str(b) and sheet_count == 1:
							storage_array[index]["sessions_total_time_minutes_phone"] = row[1]
							storage_array[index]["sessions_total_phone"] = row[2]
							storage_array[index]["sessions_avg_per_user_phone"] = row[3]
							storage_array[index]["sessions_avg_time_per_user_minutes_phone"] = row[4]
							storage_array[index]["sessions_avg_length_per_user_minutes_phone"] = row[5]
						elif "EngagementSessions" in str(b) and sheet_count == 2:
							storage_array[index]["sessions_total_time_minutes_tablet"] = row[1]
							storage_array[index]["sessions_total_tablet"] = row[2]
							storage_array[index]["sessions_avg_per_user_tablet"] = row[3]
							storage_array[index]["sessions_avg_time_per_user_minutes_tablet"] = row[4]
							storage_array[index]["sessions_avg_length_per_user_minutes_tablet"] = row[5]


						elif "InstallBaseDelta" in str(metric_type) and sheet_count == 1:
							storage_array[index]["install_base_delta"] = row[1]
							storage_array[index]["install_base_delta_percentage"] = row[2]


						elif str(metric_type) == "InstallBase" and sheet_count == 1:
							storage_array[index]["store_downloads_phone"] = row[1]
						elif str(metric_type) == "InstallBase" and sheet_count == 2:
							storage_array[index]["store_downloads_tablet"] = row[1]
						elif str(metric_type) == "InstallBase" and sheet_count == 3:
							storage_array[index]["store_downloads_total"] = row[1]


						elif "Retention" in str(b) and sheet_count == 1:
							if "d30__" in str(col_header[1]):
								storage_array[index]["retention_day_30"] = row[1]
							else:
								storage_array[index]["retention_day_0"] = row[1]
								storage_array[index]["retention_day_1"] = row[2]
								storage_array[index]["retention_day_2"] = row[3]
								storage_array[index]["retention_day_3"] = row[4]
								storage_array[index]["retention_day_4"] = row[5]
								storage_array[index]["retention_day_5"] = row[6]
								storage_array[index]["retention_day_7"] = row[7]
								storage_array[index]["retention_day_14"] = row[8]
								storage_array[index]["retention_day_30"] = row[9]

						elif "StoreDownloads" in str(b) and sheet_count == 1:
							storage_array[index]["install_base"] = row[1]

			for z in storage_array:
				new_row = [z[x] for x in z]
				write_csv.writerow(new_row)


############################################################################################################
############################################################################################################

def init_combine_api_generated_xlsx_files():
	############################
	dir_path = "0_xlsx/03_28_2023/"
	destination_path = "0_xlsx/0_combined_xlsx/" 
	combine_xlsx(dir_path, destination_path)
	############################

def init_combine_selenium_collected_xlsx_files():
	############################
	destination_path = "0_xlsx/0_combined_xlsx/" 
	dir_path = "0_selenium_app_data_apple/"
	selenium_combine_xlsx(dir_path, destination_path)
	############################

init_combine_selenium_collected_xlsx_files()